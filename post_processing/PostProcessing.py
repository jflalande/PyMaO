import os
import time
import logging  # Log info output
import datetime
import argparse  # Program argument parser
import json

from os import listdir
from os.path import isfile, join

import BooleanParser

import sys

from collections import Counter # For getting top or botton N
from openpyxl import Workbook   # For outputing results in xlsx (Excel format)
import datetime as dt

import numpy as np
import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as ticker

from math import floor, log as loga
"""
Usage:
    python3 <this_file> <json_config>

where:
    json_config: The file that holds the definitions of the rows and columns to be processed

Example of config file:

see `post_processing/examples/drebin.json`

    {
        "rows":{
            "Dataset 1":"/path/to/dataset/1",  ----> Directories that contain JSON
            "Dataset 2":"/path/to/dataset/2",        files generated by the orchestrator
            "Dataset 3":"/path/to/dataset/3",
        },
        "columns":{
            "Column 1":[
                "<boolean expression with JSONPath>",
                null        ---------------------------> No special porcentage is done
            ],
            "Column 2":[
                "<boolean expression with JSONPath>",
                null
            ],
            "Column 3":[
                "<boolean expression with JSONPath>",
                "Column 2"               ------------> Porcentage in relation to Column 2,
            ]                                          which MUST be declared before
        },
        "histograms":{
            "Data 1":[
    			"<JSONPath expression>",
    			"<type>"             ---------------> This can be date, int, float, string
    		],
    		"Data 2":[
    			"<JSONPath expression>",
    			"<type>"
    		]
        },
        "output_dir":"/path/to/your/output/dir"
    }


The JSONPath implementation is from (https://github.com/h2non/jsonpath-ng)

The boolean expression parser implementation is from (https://gist.github.com/leehsueh/1290686)

It follows this grammar:

    Expression --> Terminal (>,<,>=,<=,==) Terminal
    Terminal --> Number or String or Variable

To compare a variable to a string, they must be enclosed in <""> in the JSON file. For example:

    $..Unzip.status == \"done\"

This expression compares if the variable `$..Unzip.status` is equal to the string `done`.

For our purposes, all the JSONPath expressions are variables.

### TODO:
* [ ]  Expand expression for multiple requests
"""

# Name without extension
OUTPUT_FILENAME = "res"
# Adds a very verbose level of logs
DEBUG_LEVELV_NUM = 9
logging.addLevelName(DEBUG_LEVELV_NUM, "DEBUGV")


def debugv(self, message, *args, **kws):
    # Yes, logger takes its '*args' as 'args'.
    if self.isEnabledFor(DEBUG_LEVELV_NUM):
        self._log(DEBUG_LEVELV_NUM, message, args, **kws)


logging.Logger.debugv = debugv
log = logging.getLogger("post-processing")


# Tries to apply colors to logs
def applyColorsToLogs():
    try:
        import coloredlogs

        style = coloredlogs.DEFAULT_LEVEL_STYLES
        style['debugv'] = {'color': 'magenta'}
        coloredlogs.install(
            show_hostname=False, show_name=True,
            logger=log,
            level=DEBUG_LEVELV_NUM,
            fmt='%(asctime)s [%(levelname)8s] %(message)s'
            # Default format:
            # fmt='%(asctime)s %(hostname)s %(name)s[%(process)d] %(levelname)s
            #  %(message)s'
        )
    except ImportError:
        log.error("Can't import coloredlogs, logs may not appear correctly.")


def logSetup(level):
    # if -vv option as program argument
    if level == 2:
        log.setLevel(DEBUG_LEVELV_NUM)
        log.info("Debug is Very Verbose.")
    # if -v option as program argument
    elif level == 1:
        log.setLevel(logging.DEBUG)
        log.info("Debug is Verbose.")
    # if no (-v) option
    elif level == 0:
        log.setLevel(logging.INFO)
        log.info("Debug is Normal.")
    else:
        # else
        log.setLevel(logging.INFO)
        log.warning("Logging level \"{}\" not defined, setting \"normal\" instead"
                    .format(level))


applyColorsToLogs()

################################################################################
#                                                                              #
#                             Function definitions                             #
#                                                                              #
################################################################################


def epoch_to_date(epoch):
    return datetime.datetime.fromtimestamp(epoch).strftime('%Y-%m-%d %H:%M:%S')


os.stat_float_times(False)


def output_histograms(datasets,histograms_def,output_dir):

    # Pretty colors for the delight of the eye
    plt.style.use('seaborn-deep')

    # joint_histogram = {}
    # for histogram in histograms_def:
    #     joint_histogram[name] = {'type':histograms_def[1], 'data'=[]}

    for histogram_name in histograms_def:
        log.info("Processing histogram " + histogram_name)
        joint_histogram = {}
        joint_histogram = { 'type':histograms_def[histogram_name][1], 'data':{} }

        # rows = []

        for row_name in datasets:

            # log.debug("Assigning row " + row_name)
            log.debugv("Assigning row " + row_name)
            row = datasets[row_name]

            # Print individual histograms
            # Verify if there are histograms to output
            # if len(row.histogram_collection.keys()) != 0:
                # for histogram_name in row.histogram_collection:
                # def output_histograms(data,output_dir):
            data = row.histogram_collection[histogram_name]['data']

            log.debugv("Data size is " + str(len(data)))

            hist_type = row.histogram_collection[histogram_name]['type']

            # Collect all the data from this row according to the name of
            # the histogram
            joint_histogram['data'][row_name] = data

            log.debugv("The data of " + histogram_name + " is " + str(data))

            if hist_type == 'date':
                log.debug("Histogram " + histogram_name + " is type date")

                # Processing the dates to get the maximun and minimum month-year
                mindate = dt.datetime.fromtimestamp(min(data))
                maxdate = dt.datetime.fromtimestamp(max(data))
                bindate = dt.datetime(year=mindate.year, month=mindate.month, day=1)
                mybins = [bindate.timestamp()]
                while bindate < maxdate:
                    if bindate.month == 12:
                        bindate = dt.datetime(year=bindate.year + 1, month=1, day=1)
                    else:
                        bindate = dt.datetime(year=bindate.year, month=bindate.month + 1, day=1)
                    mybins.append(bindate.timestamp())
                mybins = mdates.epoch2num(mybins)

                plot_data = mdates.epoch2num(data)

                fig, ax = plt.subplots(1, 1, figsize=(300, 40), facecolor='white')
                # fig, ax = plt.subplots(1, 1, facecolor='white')
                ax.hist(plot_data, bins=mybins, ec='black')
                ax.set_title(histogram_name + " - " + row_name)
                ax.xaxis.set_major_locator(mdates.MonthLocator())
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%m.%y'))
                fig.autofmt_xdate()
            elif hist_type == 'int':
                log.debug("Histogram " + histogram_name + " is type int")
                fig, ax = plt.subplots(1, 1, figsize=(200, 40))
                # fig, ax = plt.subplots(1,1)
                # Ajust bins
                max_exp = int(floor(loga(max(data), 10)))
                binwidth = 10**(max_exp - 3)

                ax.hist(data, bins=np.arange(min(data), max(data) + binwidth, binwidth), ec='black')
                ax.set_title(histogram_name + " - " + row_name)
                ax.set_xlim(left=0)  # Start at left zero
                ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: format(int(x), ',')))  # The formatter for the labels in the ticks (ticks are the marks withc numbers in the x axis)
                ax.xaxis.set_major_locator(ticker.MultipleLocator(binwidth*10))
                plt.xticks(rotation=45)  # Rotate x ticks
                plt.gcf().subplots_adjust(bottom=0.15)  # Adjust the lables if they go pass the figure
                plt.grid(linestyle="--")  # Grid in the histogram
            else:
                fig, ax = plt.subplots(1, 1)
                ax.hist(data, bins='auto', ec='black')
                ax.set_title(histogram_name + " - " + row_name)

            filename = histogram_name + "_" + row_name
            # plt.savefig(output_dir + "/histogram_"+ filename + ".png")
            # log.info("Histogram saved as histogram_"+ filename + ".png")
            plt.savefig(output_dir + "/histogram_"+ filename + ".pdf")
            log.info("Histogram saved as histogram_"+ filename + ".pdf")

            plt.figure().clear()
            plt.close(plt.figure())
            # print("Figure clear")

        # Draw joint histogram

        # ax.hist(dataset_list, bins, label=['x', 'y'])
        label = []
        data = []
        mix_data = []

        for row_name in joint_histogram['data']:
            log.debug("Getting " + row_name)
            label.append(row_name)
            row_data = joint_histogram['data'][row_name]
            # print(str(row_name) + " is of type " + str(type(row_data)))
            data.append(row_data)
            mix_data.extend(row_data)

        log.debug("dataset has " + str(len(data)) + " entries")

        for num in data:
            # print("num is " + str(num))
            log.debug("The size of num is " + str(len(num)))

        s = "+"
        joint = s.join(label)

        if hist_type == 'date':
            log.debug("Histogram " + histogram_name + " is type date")
            # Processing the dates to get the maximun and minimum month-year
            mindate = dt.datetime.fromtimestamp(min(mix_data))
            maxdate = dt.datetime.fromtimestamp(max(mix_data))
            bindate = dt.datetime(year=mindate.year, month=mindate.month, day=1)
            mybins = [bindate.timestamp()]
            while bindate < maxdate:
                if bindate.month == 12:
                    bindate = dt.datetime(year=bindate.year + 1, month=1, day=1)
                else:
                    bindate = dt.datetime(year=bindate.year, month=bindate.month + 1, day=1)
                mybins.append(bindate.timestamp())
            mybins = mdates.epoch2num(mybins)

            # plot_data = mdates.epoch2num(data)
            plot_data = []
            for list in data:
                plot_data.append(mdates.epoch2num(list))

            fig, ax = plt.subplots(1,1, figsize=(200, 20), facecolor='white')
            # fig, ax = plt.subplots(1,1,facecolor='white')
            ax.hist(plot_data, bins=mybins, ec='black')
            log.debug("This title is " + histogram_name + " - " + joint)
            ax.set_title(histogram_name + " - " + joint)
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m.%y'))
            # fig.autofmt_xdate()

            # Changing the space between the ticks

            # plt.gca().margins(x=0)
            # plt.gcf().canvas.draw()
            tl = plt.gca().get_xticklabels()
            # log.debug("tl = " + str(tl))
            ticks_label = [t.get_window_extent().width for t in tl]
            log.debug(str(ticks_label))
            maxsize = max(ticks_label)

            # If the ticks label list brings 0 in everyting
            if maxsize == 0:
                maxsize = 4
            log.debug("maxsize = " + str(maxsize))
            m = 0.2 # inch margin
            N = len(mix_data)
            log.debug("N = " + str(N))
            log.debug("plt.gcf().dpi = " + str(plt.gcf().dpi))
            s = maxsize/plt.gcf().dpi*N+2*m
            margin = m/plt.gcf().get_size_inches()[0]
            #
            # print("plt.gcf().get_size_inches()[1] = " + str(plt.gcf().get_size_inches()[1]))
            log.debug("plt.gcf().get_size_inches() = " + str(plt.gcf().get_size_inches()))
            log.debug("s = " + str(s))
            plt.gcf().subplots_adjust(left=margin, right=1.-margin)
            # plt.gcf().set_size_inches(s, 10)
            plt.gcf().set_size_inches(s*0.25, plt.gcf().get_size_inches()[1])

            fig.autofmt_xdate()

        elif hist_type == 'int':
            log.debug("Histogram " + histogram_name + " is type int")
            fig, ax = plt.subplots(1, 1, figsize=(200, 20))
            # fig, ax = plt.subplots(1,1)
            plt.style.use('seaborn-deep')

            max_exp = int(floor(loga(max(mix_data),10)))
            binwidth = 10**(max_exp - 3)

            ax.hist(data, bins=np.arange(min(mix_data), max(mix_data) + binwidth, binwidth),label=label)
            ax.legend(loc='upper right')
            log.debug("This title is " + histogram_name + " - " + joint)
            ax.set_title(histogram_name + " - " + joint)
            ax.set_xlim(left=0) # Start at left zero
            ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: format(int(x), ','))) # The formatter for the labels in the ticks (ticks are the marks withc numbers in the x axis)
            ax.xaxis.set_major_locator(ticker.MultipleLocator(binwidth*10))
            plt.xticks(rotation=45) # Rotate x ticks
            plt.gcf().subplots_adjust(bottom=0.15) # Adjust the lables if they go pass the figure
            plt.grid(linestyle="--") # Grid in the histogram
        else:
            fig, ax = plt.subplots(1,1)
            ax.hist(data, bins='auto', ec='black')
            log.debug("This title is " + histogram_name + " - " + joint)
            ax.set_title(histogram_name + " - " + joint)

        filename = histogram_name + "_" + joint
        # plt.savefig(output_dir + "/histogram_"+ filename + ".png")
        # log.info("Histogram saved as histogram_"+ filename + ".png")
        plt.savefig(output_dir + "/histogram_"+ filename + ".pdf")
        log.info("Histogram saved as histogram_"+ filename + ".pdf")
        plt.figure().clear()


def output_to_files(datasets,out_dir):
    log.info("Processing finished, outputing files")

    jsonfile = out_dir + "/" + OUTPUT_FILENAME + ".json"
    xlsxfile = out_dir + "/" + OUTPUT_FILENAME + ".xlsx"
    rawfile = out_dir + "/" + OUTPUT_FILENAME + "_raw.json"

    # Dictionary for the output JSON file
    dico = {}
    # Raw results of the histogram
    raw = {}
    row_num = 1

    # Initialize a workbook
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Create names for columns in workbook
    for row_name in datasets:
        row = datasets[row_name]
        ws.cell(row=row_num, column=2, value="Total")
        col_num = 3
        for column in row.columns:
            ws.cell(row=row_num, column=col_num, value=column.name)
            ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + 1)
            col_num += 2

    row_num += 1
    col_num = 1

    # If the result JSON already exists, don't do anything
    # if not os.path.isfile(jsonfile):
    for row_name in datasets:
        log.debug("Assigning row " + row_name)

        dico[row_name] = {}
        raw[row_name] = {}

        row = datasets[row_name]

        # Assigning the name of the row
        ws.cell(row=row_num, column=col_num,value=row_name)
        col_num += 1

        # Assigning the total of the row
        ws.cell(row=row_num,column=col_num,value=row.total)
        ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num+1, end_column=col_num)
        col_num += 1

        if len(row.columns) != 0:
            for column in row.columns:
                log.debug("Assiging column " + column.name)
                column.get_total()
                if column.req_type is None:
                    dico[row_name][column.name] = [column.total, column.pct_total, column.pct_depend]
                    ws.cell(row=row_num, column=col_num, value=column.total)
                    ws.cell(row=row_num+1, column=col_num, value=column.pct_total)
                    ws.cell(row=row_num+1, column=col_num+1, value=column.pct_depend)
                    # Grand total merge_cells
                    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num+1)
                else:
                    dico[row_name][column.name] = column.res_poll
                    ws.cell(row=row_num, column=col_num, value=str(column.res_poll)
                            .replace("{", "").replace("}", "").replace(",", "\n"))
                    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num+1, end_column=col_num+1)
                col_num += 2
            # Dataset name merge
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num+1, end_column=1)
            row_num += 2
            col_num = 1
        # print(str(dico))
        if 'histograms' in row.histogram_collection.keys():
            for histogram_name in row.histogram_collection:
                # def output_histograms(data,output_dir):
                data = row.histogram_collection[histogram_name]['data']
                raw[row_name][histogram_name] = row.histogram_collection[histogram_name]

    # Create the output directory if it doesn't exists
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    with open(jsonfile, 'w') as out:
        json.dump(dico, out)
    log.info("File saved at " + jsonfile)

    with open(rawfile, 'w') as out:
        json.dump(raw, out)
    log.info("File saved at " + rawfile)

    # filename = "res.xlsx"
    wb.save(xlsxfile)
    log.info("File saved at " + xlsxfile)
    # else:
    #     log.warning("file already written. finishing")


def topN(dico, N, poll_type):
    if poll_type == 'top':
        res = dict(Counter(dico).most_common(N))
    elif poll_type == 'bottom' or poll_type == 'bot':
        res = dict(Counter(dico).most_common()[:-N-1:-1])
    return res

################################################################################
#                                                                              #
#                          Row and column definition                           #
#                                                                              #
################################################################################


class Column:
    def __init__(self, name, parent_row, req, depend=None):
        self.name = name
        self.total = self.pct_total = self.pct_depend = 0
        self.parent_row = parent_row
        self.req = req
        # self.req_parser = None
        self.poll = {}
        self.res_poll = {}
        # print("The req is " + str(self.req))
        # print("The length is " + str(len(self.req.split(":"))))
        if len(self.req.split(":")) < 2:
            # Normal request
            self.req_type = None
            self.jpath = self.req.split(" ")[0]
            self.req_parser = BooleanParser.BooleanParser(self.req)
        else:
            # This is poll request, it only needs to retrive a value
            # self.req_type can be top or bottom
            self.jpath, self.req_type, self.num_poll = self.req.split(":")
            self.num_poll = int(self.num_poll)
        self.depend = depend
        # print("Column " + self.name  + " created")

    def get_total(self):
        if self.req_type is None:
            # If there is no samples in the column, throw a warning
            if self.total == 0:
                log.warning(self.name + ": The total is 0, cannot divide by 0")
            else:
                # Calculate the percentage by total
                self.pct_total = (self.total / self.parent_row.total)*100
                # If no dependence is declared, the percentage is by total
                if self.depend is None:
                    self.pct_depend = self.pct_total
                else:
                    self.pct_depend = (self.total / self.depend.total)*100
        else:
            # print('getting poll')
            self.res_poll = topN(self.poll, self.num_poll, self.req_type)


class Row:
    var_dict = {}

    def __init__(self, name):
        self.name = name
        self.total = 0
        # self.
        self.columns = []
        self.histogram_collection = {}
        # print("Row " + self.name + " created")

    def create_column(self, name, req, depend_name=None):
        depend = None
        # If the name of the column for getting the percentage is different from None
        if depend_name is not None:
            # Check if the name exists:
            # Find the name in self.columns "column_name"
            # If the tuple is empty, the name was not found, depends is None
            column_name = (column for column in self.columns if column.name == depend_name)
            depend = next(column_name, None)
        # print('this depend is ' + str(depend))
        new_column = Column(name, self, req, depend)
        self.columns.append(new_column)

    def create_histogram(self, name, request, type):
        self.histogram_collection[name] = {"type": type, "request": request, "data": []}

    def process(self, json_file):
        self.total += 1

        if len(self.columns) != 0:
            for column in self.columns:
                log.debugv('Processing data ' + str(self.total) )
                log.debugv('parsing: ' + str(column.name))

                # Find the value of the JSONPath expression if it's not in the dictionary
                expression = column.jpath

                if expression not in self.var_dict.keys():
                    try:
                        parse = expression.split('.')
                        name = json_file['name']
                        log.debugv('the parsed list is: ' + str(parse))
                        log.debugv('My name is: ' + str(parse[2]))
                        self.var_dict[expression] = json_file[name][parse[2]][parse[3]]  # This may fail
                        if self.var_dict[expression] == "":
                            self.var_dict[expression] = 0
                    except Exception as e:
                        log.debugv("This happened during the parsing: " + str(e))
                        self.var_dict[expression] = None
                    # self.var_dict[expression] = parse(expression).find(json_file)
                log.debugv("This is the result of the parsing: " + str(self.var_dict[expression]))
                # If the expression's value is None (which it was not found in the JSON),
                # do nothing. Else, add the value
                if self.var_dict[expression] is not None:
                    # val = self.var_dict[expression][0].value
                    val = self.var_dict[expression]
                    log.debugv("The value for " + str(expression) + " is: " + str(val))
                    # Put some quotes to strings, so Jason can be happy :DDDD
                    if not isinstance(val, int) or isinstance(val, float):
                        log.debugv('There is an instance of string for val: ' + str(val))
                        val = "\"" + val + "\""
                    elif isinstance(val, bool):
                        log.debugv('There is an instance of bool for val: ' + str(val))
                        val = "\"" + str(val) + "\""
                    if column.req_type is None:
                        log.debugv('changing ' + column.jpath + ' -> ' + str(val))
                        evaluator = column.req_parser
                        log.debugv('The request to evaluate is: ' + str(evaluator.tokenizer.expression))
                        # {expression:val} => change 'expression' to 'val' when evaluating
                        res = evaluator.evaluate({expression: val})
                        log.debugv("The result was: " + str(res))
                        if res:
                            column.total += 1
                    else:
                        val = str(val)
                        if val in column.poll.keys():
                            column.poll[val] += 1
                        else:
                            column.poll[val] = 1

        if (self.histogram_collection.keys()) != 0:
            for histogram_name in self.histogram_collection:
                histogram_dict = self.histogram_collection[histogram_name]

                # Find the value of the JSONPath expression if it's not in the dictionary
                expression = histogram_dict['request']

                if expression not in self.var_dict.keys():
                    try:
                        parse = expression.split('.')
                        name = json_file['name']
                        log.debugv('the parsed list is: ' + str(parse))
                        log.debugv('My name is: ' + str(parse[2]))
                        self.var_dict[expression] = json_file[name][parse[2]][parse[3]]
                        if histogram_dict['type'] == 'date' and self.var_dict[expression] == "":
                            log.debugv('This date "' + expression + '" is now 0')
                            self.var_dict[expression] = 0
                    except Exception as e:
                        if histogram_dict['type'] == 'data':
                            self.var_dict[expression] = 0
                        else:
                            self.var_dict[expression] = None
                log.debugv("--- Histogram: This is the result of the parsing: " + str(self.var_dict[expression]))

                # Add the value if it exists
                if self.var_dict[expression] is not None:
                    val = self.var_dict[expression]
                    log.debugv("The value for " + str(expression) + " is: " + str(val))
                    if histogram_dict['type'] == 'int':
                        histogram_dict['data'].append(int(val))
                    else:
                        histogram_dict['data'].append(val)
        else:
            log.debugv("No histograms, moving on")

        # Reinitialize the dictionary
        self.var_dict = {}

################################################################################
#                                                                              #
#                  postprocessing function (primary function)                  #
#                                                                              #
################################################################################


def postprocessing(myjsonconfig, verbose=0):

    # print('This verbosity is ' + str(verbose))
    # quit()
    #
    # logSetup(verbose)
    t_start = time.time()
    filename = myjsonconfig

    # Check if the JSON config file exists
    if not os.path.isfile(filename):
        log.warning("The file " + filename + " doesn't exist. Quitting")
    else:
        with open(filename) as config:
            myjson = json.load(config)

        log.debugv("This is the config file: " + str(myjson))

        log.info("Output dir: " + myjson['output_dir'])
        jsonfile = myjson['output_dir'] + "/" + OUTPUT_FILENAME + ".json"
        # xlsxfile = myjson['output_dir'] +  "/" + OUTPUT_FILENAME + ".xlsx"
        # rawfile = myjson['output_dir'] +  "/" + OUTPUT_FILENAME + "_raw.json"

        # Check of the result file already exists
        # if os.path.isfile(jsonfile):
        #     log.warning("The output file already exist. Quitting")

        # else:

        # Initialize Row and Columns objects
        datasets={} # An empty dataset (dictionary)

        # for each row, because they are different datasets
        for row in myjson['rows']:

            log.info('Processing ' + row)

            # create row object
            datasets[row] = Row(row)

            # create columns
            if 'columns' in myjson.keys():
                for column in myjson['columns']:
                    log.debugv("the column is " + str(column))
                    # create_column(self,name,req,depends=None):
                    datasets[row].create_column(column, myjson['columns'][column][0], myjson['columns'][column][1])
            else:
                log.debugv("There are no columns in the config file, moving on")

            # Create histograms
            if 'histograms' in myjson.keys():
                for histogram in myjson['histograms']:
                    log.debugv("Adding " + histogram + " histogram")
                    # Each row will contain the same values: name, request (JSONPath), and type
                    datasets[row].create_histogram(histogram,myjson['histograms'][histogram][0],myjson['histograms'][histogram][1])
            else:
                log.debugv("There are no histograms in the config file, moving on")

            mypath = myjson['rows'][row]

            try:
                files = [f for f in listdir(mypath) if isfile(join(mypath, f)) and f.endswith(".json")]
            except Exception:
                raise Exception("Cannot open dir")

            # Process each file into the rows
            numFiles = len(files)
            log.info("Processing " + str(numFiles) + " files")
            for filename in files:

                log.debugv(row + "$ Processing file number " + str(numFiles) + ": " + filename + " JSON file")
                numFiles -= 1
                with open(mypath + "/" + filename) as f:
                    mwjson = json.load(f)

                mwjson['name'] = filename.split('.')[0]
                datasets[row].process(mwjson)

            if verbose >= 1:
                for histo in myjson['histograms']:
                    log.debug("Dataset " + row + " has " + str(len(datasets[row].histogram_collection[histo]['data'])) + " entries in " + histo)

        output_to_files(datasets, myjson['output_dir'])

        t_end = time.time()
        log.info("PROCESS TIME: " + str(round(t_end - t_start,1)) + " s")

        if 'histograms' in myjson.keys():
            if verbose == 2:
                for name in datasets:
                    log.debugv("This dataset has evaludated " + str(datasets[name].total) + " entries")
                    for histo in myjson['histograms']:
                        log.debugv("Dataset " + datasets[name].name + " has " + str(len(datasets[name].histogram_collection[histo]['data'])) + " entries in " + histo)

            output_histograms(datasets, myjson['histograms'],myjson['output_dir'])

    t_end = time.time()
    log.info("TOTAL TIME: " + str(round(t_end - t_start,1)) + " s")
    quit()


if __name__ == "__main__":
    # print("your arguments are: " + str(sys.argv))
    # Doc : https://docs.python.org/3/howto/argparse.html#id1
    parser = argparse.ArgumentParser(description="PostProcessing, for json files generated by the orchestrator.")
    parser.add_argument('json_config_file', help='The path to the JSON config file')
    parser.add_argument('-v', help='Output information to the standart output (-vv is very verbose)', action="count")
    # parser.add_argument('-H', metavar='Result JSON file', help='Output information to the standart output (-vv is very verbose)')
    args=parser.parse_args()

    if args.v is not None:
        verbosity = args.v
    else:
        verbosity = 0

    logSetup(verbosity)
    # log.info('This verbosity is ' + str(verbosity))
    # quit()
    postprocessing(args.json_config_file, verbosity)

    quit()
